"""
Web scraper for New Zealand universities.
"""

import io
import requests
import openpyxl
from typing import List, Dict

def scrape_universities(url: str) -> List[Dict[str, str]]:
    """
    Scrape universities from the New Zealand XLSX file.
    
    Args:
        url: URL to download the XLSX file
        
    Returns:
        List of dictionaries with 'name' and 'website' keys
    """
    universities = []
    
    try:
        # Create session with cookies
        session = requests.Session()
        
        # Set up browser headers
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': '*/*',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Referer': 'https://www.educationcounts.govt.nz/',
            'DNT': '1'
        }
        session.headers.update(headers)
        
        # Visit the main page first to get cookies
        session.get('https://www.educationcounts.govt.nz/directories/list-of-tertiary-providers', timeout=30)
        
        # Download the file
        response = session.get(url, timeout=30)
        response.raise_for_status()
        
        # Load it into openpyxl
        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = workbook.active
        
        # Delete the first 7 rows
        sheet.delete_rows(1, 7)
        
        # Get headers from first row
        headers_row = next(sheet.iter_rows(values_only=True))
        if not headers_row:
            print("No header row found")
            return universities
        
        # Find column indices
        headers_lower = [str(h).lower() if h else "" for h in headers_row]
        
        name_idx = None
        website_idx = None
        type_idx = None
        
        for idx, header in enumerate(headers_lower):
            if "name" in header and header == "name":
                name_idx = idx
            elif "website" in header:
                website_idx = idx
            elif header == "type":  # Exact match to avoid "Funding Type"
                type_idx = idx
        
        # Process data rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            
            # Filter: Type must be "University"
            if type_idx is not None and type_idx < len(row):
                row_type = str(row[type_idx]).strip()
                if row_type != "University":
                    continue
            
            # Extract name
            name = ""
            if name_idx is not None and name_idx < len(row):
                name = str(row[name_idx]).strip() if row[name_idx] else ""
            
            # Extract website
            website = ""
            if website_idx is not None and website_idx < len(row):
                website = str(row[website_idx]).strip() if row[website_idx] else ""
            
            if name:
                universities.append({
                    'name': name,
                    'website': website
                })
        
    except requests.exceptions.RequestException as e:
        print(f"Error downloading New Zealand file: {str(e)}")
    except Exception as e:
        print(f"Error scraping New Zealand: {str(e)}")
    
    return universities

if __name__ == "__main__":
    # Test scraper
    test_url = "https://www.educationcounts.govt.nz/__data/assets/file/0006/62574/tertiary-providers-directory-02022026.xlsx"
    results = scrape_universities(test_url)
    
    print(f"Found {len(results)} universities:")
    for uni in results:
        print(f"  - {uni['name']}: {uni['website']}")
